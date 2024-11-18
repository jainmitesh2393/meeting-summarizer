import pandas as pd
from openpyxl import Workbook
import streamlit as st
import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import speech_recognition as sr
import pyttsx3
import time
from datetime import datetime
from groq import Groq
from openpyxl.styles import Font, Alignment, PatternFill


# Initialize the recognizer and text-to-speech engine
recognizer = sr.Recognizer()
engine = pyttsx3.init()

# Function to speak text
def speak(text):
    engine.say(text)
    engine.runAndWait()

# Function to listen to user's voice and return text
def listen():
    with sr.Microphone() as source:
        recognizer.adjust_for_ambient_noise(source)
        st.info("Listening for the meeting summary... Please speak clearly.")
        audio = recognizer.listen(source)
        try:
            text = recognizer.recognize_google(audio)
            st.success(f"You said: {text}")
            return text.lower()
        except sr.UnknownValueError:
            st.warning("Sorry, I didn't catch that.")
            speak("Sorry, I didn't catch that.")
            return ""
        except sr.RequestError:
            st.error("Network error. Please check your connection.")
            speak("Network error.")
            return ""

# Function to initialize Groq API client
def initialize_groq_client():
    api_key = "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"  # Replace with your actual API key
    try:
        return Groq(api_key=api_key)
    except Exception as e:
        st.error(f"Error initializing Groq client: {e}")
        return None

# Function to generate a to-do list using the Groq API
def generate_todo_list(client, meeting_summary):
    system_prompt = """
    You are a smart assistant that generates to-do lists from meeting summaries. Your task is to read a paragraph that summarizes a meeting and extract specific action items into a concise to-do list. Each task must be formatted as follows:

    Task ID: <unique ID in ascending order>
    <task description>
    Assigned to: <person responsible for the task>
    the due date (in DD M YYYY format): <due date or "N/A" if not mentioned>
    the due time (in HHMM format): <due time or "N/A" if not mentioned>
    """

    conversation = f"Meeting Summary: {meeting_summary}\nAssistant: Please extract the tasks, create a to-do list with unique task IDs, assigned persons, due dates, and times."

    try:
        chat_completion = client.chat.completions.create(
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": conversation}
            ],
            model="llama3-70b-8192",
            temperature=0.5
        )
        response = chat_completion.choices[0].message.content
        return response.encode('utf-8').decode('utf-8')
    except Exception as e:
        st.error(f"Error generating to-do list: {e}")
        return "An error occurred while generating the to-do list."



# Function to get a brief subject from the Groq API
def extract_subject_from_summary(client, meeting_summary):
    prompt = """
    Please generate a concise subject for the following meeting summary:
    
    Summary: {meeting_summary}
    
    The subject should capture the core topic in a few words and be returned in double quotes, e.g., "Academic Planning Meeting".
    """

    conversation = prompt.format(meeting_summary=meeting_summary)

    try:
        response = client.chat.completions.create(
            messages=[
                {"role": "system", "content": "Summarize the core topic"},
                {"role": "user", "content": conversation}
            ],
            model="llama3-70b-8192",
            temperature=0.5
        )
        # Capture only the text within double quotes
        subject = response.choices[0].message.content.strip()
        if subject.startswith('"') and subject.endswith('"'):
            subject = subject[1:-1]  # Remove the double quotes
        return subject
    except Exception as e:
        st.error(f"Error extracting subject from Groq API: {e}")
        return "Meeting Summary"



# Function to send the email with an attachment
def send_email_with_attachment(to_emails, subject, body, attachment_path):
    from_email = "mitesh.22210204@viit.ac.in"  # Replace with your email
    password = "ifsl eisx hqyk eomo"  # Replace with your app password

    for email in to_emails:
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = email
        msg['Subject'] = subject

        msg.attach(MIMEText(body, 'plain'))

        # Attach the Excel file
        try:
            with open(attachment_path, "rb") as attachment:
                part = MIMEApplication(attachment.read(), Name="MoM_Output.xlsx")
                part['Content-Disposition'] = 'attachment; filename="MoM_Output.xlsx"'
                msg.attach(part)
            
            st.info("Connecting to SMTP server...")
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(from_email, password)
            st.success(f"Sending email to {email}...")

            # Send email
            server.sendmail(from_email, email, msg.as_string())
            server.quit()
            st.success(f"Email successfully sent to {email}")

        except FileNotFoundError:
            st.error(f"The file {attachment_path} was not found.")
        except smtplib.SMTPException as e:
            st.error(f"SMTP error: {e}")


# Function to display tasks in light-colored sticky notes style
def display_sticky_notes(todo_list):
    tasks = todo_list.split('\n\n')  # Split tasks by double new lines
    for task in tasks:
        if task.strip():  # Only display non-empty tasks
            st.markdown(f"""
                <div style="
                    background-color: #F8F9FA; 
                    padding: 15px; 
                    margin-bottom: 10px; 
                    border-radius: 10px; 
                    box-shadow: 2px 2px 5px rgba(0,0,0,0.1); 
                    font-family: Arial; 
                    font-size: 16px;
                    border-left: 5px solid #007BFF;
                    color: #333;
                ">
                    {task.replace('Task ID:', '<b>Task ID:</b>').replace('Assigned to:', '<b>Assigned to:</b>').replace('the due date:', '<b>Due Date:</b>').replace('the due time:', '<b>Due Time:</b>')}
                </div>
            """, unsafe_allow_html=True)

def display_mom_format(todo_list, sender_email, recipient_emails, meeting_summary):
    current_datetime = datetime.now().strftime("%d/%m/%Y, %I:%M %p")
    chairperson_name = sender_email.split('@')[0].capitalize()
    participants = ", ".join([email.split('@')[0].capitalize() for email in recipient_emails])

    # Display header for the MoM
    st.markdown(
        f"""
        <div style="background-color: #007BFF; color: white; padding: 20px; border-radius: 10px; text-align: center;">
            <h2>Minutes of Meeting</h2>
            <p style="margin: 0;">Date: {current_datetime}</p>
        </div>
        """, unsafe_allow_html=True
    )

    # Meeting Details Section
    with st.expander("Meeting Details", expanded=True):
        st.markdown(
            f"""
            *Chairperson:* {chairperson_name}  
            *Participants:* {participants}  
            *Meeting Summary:*  
            {meeting_summary}
            """, unsafe_allow_html=True
        )
    
    # Task Table
    tasks = todo_list.split('\n\n')
    st.markdown("### Action Items")
    if tasks:
        st.table(
            {
                "Sr. No.": [i + 1 for i in range(len(tasks)) if tasks[i].strip()],
                "Description": [
                    task.split('\n')[1] if len(task.split('\n')) > 1 else "N/A"
                    for task in tasks if task.strip()
                ],
                "Assigned To": [
                    task.split('\n')[2].replace("Assigned to:", "").strip()
                    if len(task.split('\n')) > 2 else "N/A"
                    for task in tasks if task.strip()
                ],
                "Due Date": [
                    task.split('\n')[3].replace("the due date:", "").strip()
                    if len(task.split('\n')) > 3 else "N/A"
                    for task in tasks if task.strip()
                ],
            }
        )
    else:
        st.warning("No action items found!")

    # Add a collapsible section for raw MoM content
    with st.expander("Generated MoM Format (Raw Text)"):
        mom_content = generate_mom_format(todo_list, sender_email, recipient_emails, meeting_summary)
        st.text_area("Minutes of Meeting Format", mom_content, height=300)

    # Button for saving or exporting the MoM
    if st.button("Export MoM as Text File"):
        with open("minutes_of_meeting.txt", "w") as file:
            file.write(mom_content)
        st.success("MoM exported successfully!")

def generate_mom_format(todo_list, sender_email, recipient_emails, meeting_summary):
    current_datetime = datetime.now().strftime("%d/%m/%Y, %I:%M %p")
    chairperson_name = sender_email.split('@')[0].capitalize()
    participants = ", ".join([email.split('@')[0].capitalize() for email in recipient_emails])

    mom_output = f"MINUTES OF MEETING\n\nDate: {current_datetime}\nVenue: [Insert Venue]\nChairperson: {chairperson_name}\nParticipants: {participants}\n\nSummary:\n{meeting_summary}\n\n"
    mom_output += "Action Items:\nSr. No. | Points Raised / Suggested | Timelines | Responsibility | Remarks\n"

    tasks = todo_list.split('\n\n')
    for idx, task in enumerate(tasks, start=1):
        if task.strip():
            lines = task.split('\n')
            task_description = lines[1] if len(lines) > 1 else "N/A"
            assigned_to = lines[2].replace("Assigned to:", "").strip() if len(lines) > 2 else "N/A"
            due_date = lines[3].replace("the due date:", "").strip() if len(lines) > 3 else "N/A"
            remarks = "Remarks: [Insert Remarks Here]"

            mom_output += f"{idx} | {task_description} | {due_date} | {assigned_to} | {remarks}\n"
    
    return mom_output


















# Function to parse MoM text into a dictionary
def parse_mom_to_dict(mom_text):
    tasks = mom_text.strip().split("\n")  # Split by lines
    parsed_data = []

    # Skip the header line
    for line in tasks[1:]:
        # Split each line by pipe (|) and trim whitespace
        columns = [col.strip() for col in line.split('|')]

        # Ensure there are at least 5 columns (Sr. No., Points, Timelines, Responsibility, Remarks)
        if len(columns) >= 5:
            parsed_data.append({
                "Sr. No.": columns[0],
                "Points Raised / Suggested": columns[1],
                "Timelines": columns[2],
                "Responsibility": columns[3],
                "Remarks": columns[4]
            })

    return parsed_data


# Function to create Excel file from MoM text with proper formatting
def mom_to_excel(mom_text, output_filename="MoM_Output.xlsx"):
    parsed_data = parse_mom_to_dict(mom_text)
    if not parsed_data:
        raise ValueError("No tasks found in MoM text to save to Excel.")

    # Create a DataFrame
    df = pd.DataFrame(parsed_data)

    # Write to Excel with formatting
    with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="MoM")
        workbook = writer.book
        worksheet = writer.sheets["MoM"]

        # Apply formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Format headers
        for col_num, column_title in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

        # Adjust column widths
        for col_num, column_cells in enumerate(worksheet.columns, start=1):
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[chr(64 + col_num)].width = max_length + 2

    return output_filename

# Function to display and save MoM format
def display_and_save_mom(todo_list, sender_email, recipient_emails, meeting_summary):
    mom_content = generate_mom_format(todo_list, sender_email, recipient_emails, meeting_summary)
    st.text_area("Minutes of Meeting Format", mom_content, height=500)

    if st.button("Save MoM to Excel"):
        try:
            output_filename = mom_to_excel(mom_content)
            st.success(f"MoM saved successfully to {output_filename}")
            with open(output_filename, "rb") as file:
                st.download_button(
                    label="Download MoM Excel File",
                    data=file,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"An error occurred while saving the MoM: {e}")









# Main function for the Streamlit app
def main():
    st.set_page_config(page_title="Meeting To-Do List Generator", page_icon="📝", layout="wide")

    # Page Header
    st.title("📝 Meeting To-Do List Generator")
    st.subheader("Easily convert your meeting summaries into actionable to-do lists")

    # Initialize session state for meeting summary
    if 'meeting_summary' not in st.session_state:
        st.session_state['meeting_summary'] = ""  # Initialize session state variable

    col1, col2 = st.columns([1, 3])

    with col1:
        st.write("## Input Options")
        use_microphone = st.radio(
            "How would you like to input your meeting summary?",
            ("Use Text", "Use Microphone")
        )

        if use_microphone == "Use Microphone":
            if st.button("🎤 Record Meeting Summary"):
                with st.spinner("Recording... Please speak into the microphone"):
                    time.sleep(1)
                    meeting_summary = listen()
                    if meeting_summary:
                        st.session_state['meeting_summary'] = meeting_summary  # Store in session state
                        st.write(f"Captured Meeting Summary: {meeting_summary}")
            else:
                st.write("Click 'Record' to capture a meeting summary.")
        else:
            # Ensure the text area works properly
            meeting_summary = st.text_area(
                "Enter your meeting summary:", 
                st.session_state.get('meeting_summary', ''),
                height=150
            )
            if st.button("Save Summary"):
                st.session_state['meeting_summary'] = meeting_summary  # Store in session state


    with col2:
        st.write("## To-Do List Output")
        if st.button("📝 Generate To-Do List"):
            meeting_summary = st.session_state.get('meeting_summary', '')
            if not meeting_summary.strip():
                st.error("Please enter or record a valid meeting summary.")
                return

            # Initialize Groq API client
            client_groq = initialize_groq_client()
            if client_groq is None:
                st.error("Failed to initialize the Groq client. Please check your API key.")
                st.stop()

            # Generate To-Do list from summary
            with st.spinner("Generating To-Do List..."):
                todo_list = generate_todo_list(client_groq, meeting_summary)
                if todo_list:
                    st.success("To-Do List Generated!")
                    display_sticky_notes(todo_list)  # Display each task in a sticky note style

                    # Send the to-do list via email

                else:
                    st.error("Failed to generate the to-do list.")
        
        # Main application logic
        sender_email = "mitesh.22210204@viit.ac.in"
        recipient_emails = ["mitesh.22210204@viit.ac.in"]  # Use the actual recipient emails

        if st.button("Generate MoM Format"):
            meeting_summary = st.session_state.get('meeting_summary', '')
            if not meeting_summary.strip():
                st.error("Please enter or record a valid meeting summary.")
            else:
                client_groq = initialize_groq_client()
                todo_list = generate_todo_list(client_groq, meeting_summary)
                display_mom_format(todo_list, sender_email, recipient_emails, meeting_summary)



        
        # Initialize Groq API client
        client_groq_1 = initialize_groq_client()
        if client_groq_1 is None:
            st.error("Failed to initialize the Groq client. Please check your API key.")
            st.stop()
        list = generate_todo_list(client_groq_1, meeting_summary)


        if st.button("📧 Send MoM via Email"):
            default_emails = ["mitesh.22210204@viit.ac.in"]
            
            # Generate the MoM content from the to-do list
            client_groq_2= initialize_groq_client()
            todo_list_1= generate_todo_list(client_groq_2, meeting_summary)
            mom_content = generate_mom_format(todo_list_1,sender_email, recipient_emails, meeting_summary)
            excel_file = mom_to_excel(mom_content, "MoM_Output.xlsx")  # Convert MoM text to Excel
            send_email_with_attachment(default_emails, "Meeting MoM", "Please find the MoM attached.", excel_file)


if __name__ == "__main__":
    main()